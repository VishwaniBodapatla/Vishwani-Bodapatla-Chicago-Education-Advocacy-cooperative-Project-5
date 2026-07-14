# pipelinescrapper_mod.py
from __future__ import annotations

import time
import re
import json
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Callable

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
OUTPUT_DIR = PROJECT_DIR / "output"

JOBRIGHT_URL = "https://jobright.ai/jobs/cybersecurity-jobs-in-united-states"

SCROLL_CONTAINER = "#scrollableDiv"
JOB_CARD_DIV_SELECTOR = "div.index_job-card__oqX1M[id]"

DETAIL_CLOSE_BTN = 'button:has(img[alt="close detail"])'

OUT_DIR = OUTPUT_DIR / "dom_snapshots"
MERGED_HTML = OUT_DIR / "container_dom_MERGED.html"
OUT_EXCEL = OUTPUT_DIR / "jobright_jobs.xlsx"
OUT_JSONL = OUT_DIR / "jobs_raw.jsonl"

STATE_JSON = OUT_DIR / "state.json"

SAVE_EVERY_N_NEW = 10
SAVE_EVERY_SECONDS = 30

FETCH_DETAIL_ON_CLICK = True
MAX_DETAIL_CLICKS: Optional[int] = None

HEADERS = [
    "platform",
    "role_selected",
    "role_name",
    "role_type",
    "work_model",
    "company_name",
    "company_url",
    "job_url",
    "salary",
    "salary_min",
    "salary_max",
    "degree_level",
    "branch",
    "location",
    "posted",
    "seniority",
]


# =========================
# Helpers
# =========================
def norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def extract_role_selected_from_url(url: str) -> str:
    m = re.search(r"/jobs/([a-z0-9-]+)-jobs-in-", (url or "").lower())
    return m.group(1).replace("-", " ") if m else ""


def autosize_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            widths[i] = max(widths.get(i, 0), len(str(v or "")))
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w + 2, 12), 55)


def parse_salary_range(s: str) -> Tuple[str, Optional[int], Optional[int]]:
    raw = norm_space(s)
    if not raw:
        return "", None, None

    nums = re.findall(r"(\d{1,3}(?:,\d{3})*|\d+)\s*([KkMm])?", raw)
    if not nums:
        return raw, None, None

    def to_int(num_str: str, suffix: str) -> int:
        n = int(num_str.replace(",", ""))
        suf = (suffix or "").lower()
        if suf == "k":
            return n * 1000
        if suf == "m":
            return n * 1_000_000
        return n

    values = [to_int(n, suf) for n, suf in nums]
    if len(values) == 1:
        return raw, values[0], None
    lo, hi = min(values[0], values[1]), max(values[0], values[1])
    return raw, lo, hi


_REAL_INFO_RE = re.compile(r"^https://jobright\.ai/jobs/info/[0-9a-f]{24}([?#].*)?$", re.I)
_REAL_SIMPLE_RE = re.compile(r"^https://jobright\.ai/jobs/[0-9a-f]{24}([?#].*)?$", re.I)


def is_real_jobright_job_url(url: str) -> bool:
    u = (url or "").strip()
    if not u.startswith("https://jobright.ai/jobs/"):
        return False
    if "/jobs/recommend" in u:
        return False
    if "-jobs-in-" in u:
        return False
    if "/jobs/search" in u:
        return False
    if _REAL_INFO_RE.match(u):
        return True
    if _REAL_SIMPLE_RE.match(u):
        return True
    return False


@dataclass
class JobRow:
    platform: str
    role_selected: str
    role_name: str
    role_type: str
    work_model: str
    company_name: str
    company_url: str
    job_url: str
    salary: str
    salary_min: Optional[int]
    salary_max: Optional[int]
    degree_level: str
    branch: str
    location: str
    posted: str
    seniority: str


def find_text_by_icon(card_soup, icon_keyword: str) -> str:
    img = card_soup.select_one(f'img[src*="{icon_keyword}"]')
    if not img:
        return ""
    wrapper = img.find_parent()
    if not wrapper:
        return ""
    span = wrapper.find("span")
    return norm_space(span.get_text(" ", strip=True) if span else "")


def try_find_job_url(card_soup) -> str:
    a = card_soup.select_one('a[href*="/jobs/info/"], a[href*="/jobs/"]')
    if not a:
        return ""
    href = (a.get("href") or "").strip()
    if href.startswith("/"):
        href = "https://jobright.ai" + href
    return href if is_real_jobright_job_url(href) else ""


def try_find_company_url(card_soup) -> str:
    a = card_soup.select_one('a[href*="/company"], a[href*="/companies"]')
    if not a:
        return ""
    href = (a.get("href") or "").strip()
    if href.startswith("/"):
        return "https://jobright.ai" + href
    return href


def card_html_to_jobrow(card_html: str, platform: str, role_selected: str) -> JobRow:
    soup = BeautifulSoup(card_html, "html.parser")

    role_name = ""
    h2 = soup.select_one("h2.index_job-title__Riiip") or soup.find("h2")
    if h2:
        role_name = norm_space(h2.get_text(" ", strip=True))

    company_name = ""
    cn = soup.select_one("div.index_company-name__jnxCX")
    if cn:
        company_name = norm_space(cn.get_text(" ", strip=True))

    posted = ""
    pt = soup.select_one("span.index_publish-time__iYAbR")
    if pt:
        posted = norm_space(pt.get_text(" ", strip=True))

    location = find_text_by_icon(soup, "location.svg")
    role_type = find_text_by_icon(soup, "time.svg")
    salary_text = find_text_by_icon(soup, "money.svg")
    work_model = find_text_by_icon(soup, "remote.svg")
    seniority = find_text_by_icon(soup, "level.svg")

    salary_raw, salary_min, salary_max = parse_salary_range(salary_text)

    return JobRow(
        platform=platform,
        role_selected=role_selected,
        role_name=role_name,
        role_type=role_type,
        work_model=work_model,
        company_name=company_name,
        company_url=try_find_company_url(soup),
        job_url=try_find_job_url(soup),
        salary=salary_raw,
        salary_min=salary_min,
        salary_max=salary_max,
        degree_level="",
        branch="",
        location=location,
        posted=posted,
        seniority=seniority,
    )


def write_merged_html(path: Path, cards_by_id: Dict[str, str]) -> None:
    body = "\n".join(cards_by_id.values())
    html = f"""<!doctype html>
<html>
<head><meta charset="utf-8"><title>Jobright Merged Cards</title></head>
<body>
<div id="merged_cards">
{body}
</div>
</body>
</html>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8-sig")


def write_excel(path: Path, rows: List[JobRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "jobs"

    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        d = asdict(r)
        ws.append([d.get(h, "") if d.get(h) is not None else "" for h in HEADERS])

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    wb.save(path)


def append_jsonl(path: Path, obj: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    line = json.dumps(obj, ensure_ascii=False)
    with path.open("a", encoding="utf-8", newline="\n") as f:
        f.write(line + "\n")
        f.flush()


def load_rows_from_jsonl(path: Path) -> Dict[str, JobRow]:
    rows: Dict[str, JobRow] = {}
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            cid = (obj.get("id") or "").strip()
            if not cid:
                continue
            obj2 = dict(obj)
            obj2.pop("id", None)
            try:
                rows[cid] = JobRow(**obj2)
            except TypeError:
                continue
    return rows


def load_state() -> dict:
    if not STATE_JSON.exists():
        return {}
    try:
        return json.loads(STATE_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_state(state: dict) -> None:
    STATE_JSON.parent.mkdir(parents=True, exist_ok=True)
    STATE_JSON.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def extract_degree_and_branch(detail_text: str) -> tuple[str, str]:
    text = norm_space(detail_text)
    tl = text.lower()

    degree_level = ""
    if re.search(r"\b(ph\.?d|doctorate)\b", tl):
        degree_level = "PhD"
    elif re.search(r"\bmaster'?s\b|\bms\b|\bm\.s\.\b", tl):
        degree_level = "Master's"
    elif re.search(r"\bbachelor'?s\b|\bbs\b|\bb\.s\.\b", tl):
        degree_level = "Bachelor's"
    elif "associate" in tl:
        degree_level = "Associate"

    branch = ""

    patterns = [
        r"""(?ix)
        (?:master(?:'s)?\s+degree\s+or\s+ph\.?d|ph\.?d\s+or\s+master(?:'s)?\s+degree|
           bachelor(?:'s)?\s+degree\s+or\s+master(?:'s)?\s+degree|
           master(?:'s)?\s+degree|
           bachelor(?:'s)?\s+degree|
           ph\.?d|doctorate)
        \s*[:\-]\s*
        (?P<branch>[^;\n\.]{5,260})
        """,
        r"""(?ix)
        degree\s*[:\-]\s*(?P<branch>[^;\n\.]{5,260})
        """,
        r"""(?ix)
        (?:bachelor(?:'s)?|master(?:'s)?|ph\.?d|doctorate)
        \s+degree\s+in\s+(?P<branch>[^;\n\.]{5,260})
        """,
        r"""(?ix)
        (?:degree\s+in|background\s+in|major\s+in|study\s+in)\s+(?P<branch>[^;\n\.]{5,260})
        """,
    ]

    for pat in patterns:
        m = re.search(pat, text)
        if m:
            branch = norm_space(m.group("branch"))
            break

    if branch:
        branch = re.sub(r"\s*,?\s*\bor\s+equivalent\s+experience\b.*$", "", branch, flags=re.I)
        branch = re.sub(r"\s*,?\s*\bor\s+\d+\+?\s+years?\b.*$", "", branch, flags=re.I)
        branch = re.sub(r"\s*,?\s*\bwith\s+\d+\+?\s+years?\b.*$", "", branch, flags=re.I)
        branch = re.sub(r"\s*\((?:preferred|required|desired|a plus)\)\s*$", "", branch, flags=re.I)
        branch = branch.strip(" ,;:-")

    return degree_level, branch


def get_job_url_from_detail_panel(page) -> str:
    href = (page.evaluate("() => window.location.href") or "").strip()
    if is_real_jobright_job_url(href):
        return href

    url2 = page.evaluate(
        """() => {
            function isReal(u) {
                if (!u) return false;
                u = String(u).trim();
                if (!u.startsWith("https://jobright.ai/jobs/")) return false;
                if (u.includes("/jobs/recommend")) return false;
                if (u.includes("-jobs-in-")) return false;
                if (u.includes("/jobs/search")) return false;
                if (/^https:\\/\\/jobright\\.ai\\/jobs\\/info\\/[0-9a-f]{24}([?#].*)?$/i.test(u)) return true;
                if (/^https:\\/\\/jobright\\.ai\\/jobs\\/[0-9a-f]{24}([?#].*)?$/i.test(u)) return true;
                return false;
            }

            const closeBtn =
                document.querySelector('button[id^="index_job-detail-close-button"]') ||
                document.querySelector('button:has(img[alt="close detail"])');

            let node = closeBtn ? closeBtn.parentElement : document.body;

            for (let depth = 0; depth < 12 && node; depth++) {
                const anchors = Array.from(node.querySelectorAll('a[href*="/jobs/"]'));
                for (const a of anchors) {
                    const href = a.href || a.getAttribute("href") || "";
                    if (isReal(href)) return href;
                }
                node = node.parentElement;
            }

            const all = Array.from(document.querySelectorAll('a[href*="/jobs/"]'));
            for (const a of all) {
                const href = a.href || a.getAttribute("href") || "";
                if (isReal(href)) return href;
            }
            return "";
        }"""
    ) or ""
    return (url2 or "").strip()


def extract_detail_panel_fields(page) -> tuple[str, str, str]:
    job_url = get_job_url_from_detail_panel(page)
    visible_text = page.inner_text("body")
    degree_level, branch = extract_degree_and_branch(visible_text)
    return job_url, degree_level, branch


def open_detail_and_capture(page, card_css: str, timeout_ms: int = 15000) -> tuple[str, str, str]:
    scroll_top = page.eval_on_selector(SCROLL_CONTAINER, "(el) => el.scrollTop")

    page.wait_for_selector(card_css, timeout=timeout_ms)
    page.eval_on_selector(card_css, "(el) => el.scrollIntoView({block:'center'})")
    page.wait_for_timeout(120)

    page.click(card_css, timeout=timeout_ms)

    try:
        page.wait_for_selector(DETAIL_CLOSE_BTN, timeout=timeout_ms)
    except PlaywrightTimeoutError:
        page.eval_on_selector(SCROLL_CONTAINER, "(el, v) => { el.scrollTop = v; }", scroll_top)
        return "", "", ""

    page.wait_for_timeout(200)

    job_url, degree_level, branch = extract_detail_panel_fields(page)

    try:
        page.click(DETAIL_CLOSE_BTN, timeout=2000)
        page.wait_for_timeout(150)
    except Exception:
        try:
            page.click('button[id^="index_job-detail-close-button"]', timeout=2000)
            page.wait_for_timeout(150)
        except Exception:
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(150)
            except Exception:
                pass

    page.eval_on_selector(SCROLL_CONTAINER, "(el, v) => { el.scrollTop = v; }", scroll_top)
    page.wait_for_timeout(80)

    return job_url, degree_level, branch


def ensure_scroll_container(page) -> None:
    page.wait_for_selector(SCROLL_CONTAINER, timeout=30000)


def get_visible_job_ids(page) -> set[str]:
    return set(
        page.evaluate(
            """() => Array.from(document.querySelectorAll("div.index_job-card__oqX1M[id]"))
                    .map(el => el.id).filter(Boolean)"""
        )
    )


def scroll_container_by_one_page(page) -> None:
    page.eval_on_selector(
        SCROLL_CONTAINER,
        """(el) => {
            const delta = Math.floor(el.clientHeight * 0.9);
            el.scrollTop = el.scrollTop + delta;
        }""",
    )


def wait_for_visible_ids_change(page, prev_ids: set[str], timeout_ms: int = 15000) -> set[str]:
    start = time.time()
    while True:
        current = get_visible_job_ids(page)
        if current and current != prev_ids:
            return current
        if (time.time() - start) * 1000 > timeout_ms:
            return current
        page.wait_for_timeout(250)


def extract_visible_cards_html(page) -> list[dict]:
    return page.evaluate(
        """() => {
            const nodes = Array.from(document.querySelectorAll("div.index_job-card__oqX1M[id]"));
            return nodes.map(n => ({ id: n.id, html: n.outerHTML }));
        }"""
    )


def get_scrolltop(page) -> int:
    try:
        return int(page.eval_on_selector(SCROLL_CONTAINER, "(el) => el.scrollTop"))
    except Exception:
        return 0


def set_scrolltop(page, value: int) -> None:
    try:
        page.eval_on_selector(SCROLL_CONTAINER, "(el, v) => { el.scrollTop = v; }", int(value))
        page.wait_for_timeout(250)
    except Exception:
        pass


def run_pipeline(
    target_jobs: int,
    headless: bool = False,
    wait_for_login: Optional[Callable[[], None]] = None,     # ✅ NEW
    on_status: Optional[Callable[[str], None]] = None,       # ✅ NEW
) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    def status(msg: str) -> None:
        print(msg)
        if on_status:
            try:
                on_status(msg)
            except Exception:
                pass

    platform = "jobright"
    role_selected = extract_role_selected_from_url(JOBRIGHT_URL) or "cybersecurity"

    cards_by_id: Dict[str, str] = {}
    rows_by_id: Dict[str, JobRow] = {}
    detail_clicks = 0

    state = load_state()
    if (not OUT_JSONL.exists()) and (not state):
        status(f"[RESUME] No prior state found at: {OUT_DIR.resolve()} (starting fresh)")

    if OUT_JSONL.exists():
        rows_by_id = load_rows_from_jsonl(OUT_JSONL)
        status(f"[RESUME] Loaded {len(rows_by_id)} rows from {OUT_JSONL}")

    if state:
        detail_clicks = int(state.get("detail_clicks", 0))
        status(f"[RESUME] detail_clicks={detail_clicks} saved_scrollTop={state.get('scrollTop', 0)}")

    if rows_by_id:
        write_excel(OUT_EXCEL, list(rows_by_id.values()))
        status(f"[AUTOSAVE] Excel refreshed: {OUT_EXCEL.resolve()}")

    last_save_ts = time.time()
    new_since_save = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context()
        page = context.new_page()

        status(f"[INFO] Opening: {JOBRIGHT_URL}")
        page.goto(JOBRIGHT_URL, wait_until="domcontentloaded", timeout=60000)

        # ✅ NEW: GUI-controlled pause (no input())
        if wait_for_login:
            status("[ACTION] Please login in the opened browser, then click 'Continue after login' in the GUI.")
            wait_for_login()
        else:
            status("[WARN] No wait_for_login callback provided. Proceeding immediately (may fail if not logged in).")

        ensure_scroll_container(page)

        try:
            page.wait_for_selector(JOB_CARD_DIV_SELECTOR, timeout=30000)
        except PlaywrightTimeoutError:
            browser.close()
            raise RuntimeError("Could not find job cards after login. Confirm JOB_CARD_DIV_SELECTOR.")

        saved_scroll = int(state.get("scrollTop", 0)) if state else 0
        if saved_scroll > 0:
            set_scrolltop(page, saved_scroll)
            status(f"[RESUME] Restored scrollTop={saved_scroll}")

        collected_ids = set(rows_by_id.keys())
        prev_visible_ids = get_visible_job_ids(page)

        def autosave(reason: str) -> None:
            nonlocal last_save_ts, new_since_save
            try:
                write_excel(OUT_EXCEL, list(rows_by_id.values()))
                write_merged_html(MERGED_HTML, cards_by_id)
                save_state({
                    "url": JOBRIGHT_URL,
                    "target_jobs": target_jobs,
                    "saved_count": len(rows_by_id),
                    "detail_clicks": detail_clicks,
                    "scrollTop": get_scrolltop(page),
                    "ts": int(time.time()),
                })
                last_save_ts = time.time()
                new_since_save = 0
                status(f"[AUTOSAVE] {reason} | rows={len(rows_by_id)} scrollTop={get_scrolltop(page)}")
            except Exception as e:
                status(f"[WARN] autosave failed: {e}")

        def ingest_visible_cards() -> None:
            nonlocal cards_by_id, rows_by_id, detail_clicks, last_save_ts, new_since_save

            for item in extract_visible_cards_html(page):
                cid = (item.get("id") or "").strip()
                html = item.get("html") or ""
                if not cid or not html:
                    continue
                if cid in rows_by_id:
                    continue

                cards_by_id[cid] = html
                row = card_html_to_jobrow(html, platform=platform, role_selected=role_selected)

                if FETCH_DETAIL_ON_CLICK:
                    if MAX_DETAIL_CLICKS is None or detail_clicks < MAX_DETAIL_CLICKS:
                        need = (not row.job_url) or (not row.degree_level) or (not row.branch)
                        if need:
                            card_css = f'{JOB_CARD_DIV_SELECTOR}[id="{cid}"]'
                            detail_url, degree_level, branch = open_detail_and_capture(page, card_css)
                            detail_clicks += 1

                            if detail_url and not row.job_url:
                                row.job_url = detail_url
                            if degree_level and not row.degree_level:
                                row.degree_level = degree_level
                            if branch and not row.branch:
                                row.branch = branch

                rows_by_id[cid] = row
                append_jsonl(OUT_JSONL, {"id": cid, **asdict(row)})
                collected_ids.add(cid)

                new_since_save += 1
                if new_since_save >= SAVE_EVERY_N_NEW:
                    autosave("batch")
                elif (time.time() - last_save_ts) >= SAVE_EVERY_SECONDS:
                    autosave("timer")

        collected_ids |= prev_visible_ids
        ingest_visible_cards()
        autosave("startup")

        no_growth = 0

        try:
            while len(collected_ids) < target_jobs:
                before_ids = len(collected_ids)
                before_cards = len(cards_by_id)

                scroll_container_by_one_page(page)
                page.wait_for_timeout(250)

                new_visible_ids = wait_for_visible_ids_change(page, prev_visible_ids, timeout_ms=15000)
                if new_visible_ids:
                    collected_ids |= new_visible_ids

                ingest_visible_cards()

                write_merged_html(MERGED_HTML, cards_by_id)

                after_ids = len(collected_ids)
                after_cards = len(cards_by_id)
                status(f"[STEP] ids={after_ids} cards={after_cards} (+{after_cards - before_cards})")

                if after_ids <= before_ids and after_cards == before_cards:
                    no_growth += 1
                else:
                    no_growth = 0

                if no_growth >= 1500:
                    status("[STOP] No growth for many scrolls.")
                    break

                prev_visible_ids = new_visible_ids

            autosave("finished")
        except KeyboardInterrupt:
            status("\n[INTERRUPT] CTRL+C received. Saving progress...")
            autosave("keyboard interrupt")
        except Exception as e:
            status(f"\n[CRASH] {e}\nSaving progress...")
            autosave("exception crash")
            raise
        finally:
            browser.close()

    write_excel(OUT_EXCEL, list(rows_by_id.values()))
    status(f"[DONE] Excel: {OUT_EXCEL.resolve()}")
    status(f"[DONE] Merged HTML: {MERGED_HTML.resolve()}")
    status(f"[DONE] Debug JSONL: {OUT_JSONL.resolve()}")
    status(f"[DONE] Resume state: {STATE_JSON.resolve()}")